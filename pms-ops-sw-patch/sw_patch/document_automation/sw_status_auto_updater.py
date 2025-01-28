import openpyxl
import openpyxl.utils
import pandas as pd


def copy_cell_value(src_ws, src_row, src_col, dest_ws, dest_row, dest_col):
    value = src_ws.cell(row=src_row, column=src_col).value
    dest_ws.cell(row=dest_row, column=dest_col, value=value)


def log_copy_operation(
    cell_value, row_in_ws1, row_in_ws4, src_ws, src_row, src_col, target_col
):
    print(
        f"동일한 값 발견: {cell_value} (SW현황의 {row_in_ws1}행, 최신 검증서의 {row_in_ws4}헹)"
    )
    print(
        f"최신 검증서의 {openpyxl.utils.get_column_letter(src_col)}열 값 '{src_ws.cell(row=src_row, column=src_col).value}'을 SW현황의 {openpyxl.utils.get_column_letter(target_col)}열로 복사"
    )


# "SW 현황 및 검증서" 불러오기
wb = openpyxl.local_workbook("SW_현황_및_검증서".xlsx)

# SW현황과 최신검증서 가져오기
ws1 = wb.worksheets[0]  # SW현황
ws4 = wb.worksheets[3]  # 최신 검증서

# SW현황의 D열(제품명) 데이터를 딕셔너리로 저장
d_column_values = {cell.value for cell in ws1["D"] if cell.value is not None}

# 최신 검증서의 E열(제품명) 데이터와 비교하여 동일한 값이 있는 행의 데이터를 SW현황으로 복사
for cell in ws4["E"]:
    if cell.value in d_column_values:
        # 동일한 값이 있는 행 번호
        row_in_ws1 = d_column_values[cell.value]
        row_in_ws4 = cell.row

        # 해당하는 셀 복사 작업
        copy_cell_value(ws4, row_in_ws4, 4, ws1, row_in_ws1, 6)  # D열(제품명) 값 복사
        log_copy_operation(cell.value, row_in_ws1, row_in_ws4, ws4, row_in_ws4, 4, 6)
        copy_cell_value(ws4, row_in_ws4, 8, ws1, row_in_ws1, 7)  # H열(버전) 값 복사
        log_copy_operation(cell.value, row_in_ws1, row_in_ws4, ws4, row_in_ws4, 8, 7)
        copy_cell_value(
            ws4, row_in_ws4, 9, ws1, row_in_ws1, 8
        )  # I열(패치 파일명) 값 복사
        log_copy_operation(cell.value, row_in_ws1, row_in_ws4, ws4, row_in_ws4, 9, 8)
# 엑셀 파일 저장 및 닫기
wb.save("SW_현황_및_검증서.xlsx")
wb.close()
